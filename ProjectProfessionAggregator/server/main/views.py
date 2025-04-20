from django.shortcuts import render
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import authenticate, login, logout
from .serializers import RegisterSerializer, LoginSerializer, UserProfileSerializer, UserSerializer, AddVacanciesSerializer, CompanySerializer
from .models import UserProfile, Company, Vacancies, User, VacanciesMember
from rest_framework.permissions import IsAuthenticated
from .permissions import IsAdminOrReadOnly, IsAuthenticated, IsEmployerOwner, IsSuperAdmin, IsEmployerOrAdminOrSuperAdmin
from django.db import IntegrityError
from rest_framework_simplejwt.tokens import RefreshToken
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font


class RegisterAPI(APIView):
    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            try:
                user = serializer.save()
                user_auth = authenticate(request, username=user.username, password=request.data['password'])
                if user_auth:
                    login(request, user_auth)
                    refresh = RefreshToken.for_user(user)
                    return Response({
                        'refresh': str(refresh),
                        'access': str(refresh.access_token),
                        'user': {
                            'username': user.username,
                            'email': user.email
                        }
                    }, status=status.HTTP_201_CREATED)
                return Response({'error': 'Authentication failed'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginAPI(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            username = serializer.validated_data['username']
            password = serializer.validated_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                refresh = RefreshToken.for_user(user)
                return Response({
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                    'user': {
                        'username': user.username,
                        'email': user.email
                    }
                }, status=status.HTTP_200_OK)
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LogoutAPI(APIView):
    def post(self, request):
        if request.user.is_authenticated:
            logout(request)
            return Response({'message': 'Successfully logged out'}, status=status.HTTP_200_OK)
        return Response({'error': 'Not authenticated'}, status=status.HTTP_401_UNAUTHORIZED)


class ProfileAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        profile, created = UserProfile.objects.get_or_create(
            user=request.user,
            defaults={
                'name': '',
                'surname': '',
                'patronymic': '',
                'phone': '',
                'about': '',
                'years': '',
                'experience': '',
                'admin_type': 'none'
            }
        )
        serializer = UserSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def put(self, request):
        user = request.user
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={
                'name': '',
                'surname': '',
                'patronymic': '',
                'phone': '',
                'about': '',
                'years': '',
                'experience': ''
            }
        )
        user_data = {
            'username': request.data.get('username', user.username),
            'email': request.data.get('email', user.email),
        }
        user_serializer = UserSerializer(user, data=user_data, partial=True)
        if not user_serializer.is_valid():
            return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        profile_serializer = UserProfileSerializer(profile, data=request.data, partial=True)
        if not profile_serializer.is_valid():
            return Response(profile_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        password = request.data.get('password')
        if password:
            user.set_password(password)
        user_serializer.save()
        profile_serializer.save()
        user.save()
        return Response(
            {
                'user': user_serializer.data,
                'profile': profile_serializer.data,
            },
            status=status.HTTP_200_OK
        )


class UserManagementAPI(APIView):
    permission_classes = [IsSuperAdmin]
    def get(self, request, user_id=None):
        try:
            if user_id:
                user = User.objects.get(id=user_id)
                profile = user.userprofile
                data = {
                    'id': user.id,
                    'username': user.username,
                    'is_staff': user.is_staff,
                    'admin_type': profile.admin_type,
                }
                return Response(data, status=status.HTTP_200_OK)
            users = User.objects.all()
            data = []
            for user in users:
                try:
                    profile = user.userprofile
                    data.append({
                        'id': user.id,
                        'username': user.username,
                        'is_staff': user.is_staff,
                        'admin_type': profile.admin_type,
                    })
                except UserProfile.DoesNotExist:
                    data.append({
                        'id': user.id,
                        'username': user.username,
                        'is_staff': user.is_staff,
                        'admin_type': 'none',
                    })
            return Response(data, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            profile, created = UserProfile.objects.get_or_create(
                user=user,
                defaults={'admin_type': 'none'}
            )
            admin_type = request.data.get('admin_type')
            is_staff = request.data.get('is_staff', user.is_staff)

            if admin_type not in ['super', 'employer', 'none']:
                return Response({'error': 'Недопустимый тип администратора'}, status=status.HTTP_400_BAD_REQUEST)

            profile.admin_type = admin_type
            user.is_staff = is_staff and admin_type != 'none'
            profile.save()
            user.save()

            return Response({
                'id': user.id,
                'username': user.username,
                'is_staff': user.is_staff,
                'admin_type': profile.admin_type,
            }, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class CompanyAPI(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAdminOrReadOnly()]
        return [IsAuthenticated(), IsEmployerOrAdminOrSuperAdmin()]
    def get(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
            if not (company.owner == request.user or request.user.is_superuser):
                return Response(
                    {'error': 'У вас нет прав для просмотра этой компании'},
                    status=status.HTTP_403_FORBIDDEN
                )
            serializer = CompanySerializer(company)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Company.DoesNotExist:
            return Response({'error': 'Компания не найдена'}, status=status.HTTP_404_NOT_FOUND)
    def post(self, request):
        serializer = CompanySerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
            if not (company.owner == request.user or request.user.is_superuser):
                return Response(
                    {'error': 'У вас нет прав для редактирования этой компании'},
                    status=status.HTTP_403_FORBIDDEN
                )
            serializer = CompanySerializer(company, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Company.DoesNotExist:
            return Response({'error': 'Компания не найдена'}, status=status.HTTP_404_NOT_FOUND)
    def delete(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
            if not (company.owner == request.user or request.user.is_superuser):
                return Response(
                    {'error': 'У вас нет прав для удаления этой компании'},
                    status=status.HTTP_403_FORBIDDEN
                )
            company.delete()
            return Response(
                {'message': 'Компания успешно удалена'},
                status=status.HTTP_204_NO_CONTENT
            )
        except Company.DoesNotExist:
            return Response({'error': 'Компания не найдена'}, status=status.HTTP_404_NOT_FOUND)


class CompaniesListAPI(APIView):
    permission_classes = [IsAdminOrReadOnly]
    def get(self, request):
        try:
            companies = Company.objects.all()
            serializer = CompanySerializer(companies, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AddVacanciesAPI(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAdminOrReadOnly()]
        return [IsAuthenticated(), IsEmployerOrAdminOrSuperAdmin()]
    def get(self, request, id):
        try:
            vacancy = Vacancies.objects.get(id=id)
            serializer = AddVacanciesSerializer(vacancy)
            data = serializer.data
            members = VacanciesMember.objects.filter(vacancy=vacancy).select_related('user')
            data['members'] = [
                {'id': member.user.id, 'username': member.user.username, 'email': member.user.email}
                for member in members
            ]
            return Response(data, status=status.HTTP_200_OK)
        except Vacancies.DoesNotExist:
            return Response({'error': 'Вакансия не найдена'}, status=status.HTTP_404_NOT_FOUND)
    def post(self, request):
        serializer = AddVacanciesSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                vacancy = serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request, id):
        try:
            vacancy = Vacancies.objects.get(id=id)
            if not (vacancy.owner == request.user or request.user.is_superuser):
                return Response({'error': 'У вас нет прав для редактирования этой вакансии'}, status=status.HTTP_403_FORBIDDEN)
            self.check_object_permissions(request, vacancy)
            serializer = AddVacanciesSerializer(vacancy, data=request.data, partial=True)
            if serializer.is_valid():
                vacancy = serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Vacancies.DoesNotExist:
            return Response({'error': 'Вакансия не найдена'}, status=status.HTTP_404_NOT_FOUND)
    def delete(self, request, id):
        try:
            vacancy = Vacancies.objects.get(id=id)
            if not (vacancy.owner == request.user or request.user.is_superuser):
                return Response({'error': 'У вас нет прав для редактирования этой компании'}, status=status.HTTP_403_FORBIDDEN)
            self.check_object_permissions(request, vacancy)
            vacancy.delete()
            return Response({'message': 'Вакансия успешно удалена'}, status=status.HTTP_204_NO_CONTENT)
        except Vacancies.DoesNotExist:
            return Response({'error': 'Вакансия не найдена'}, status=status.HTTP_404_NOT_FOUND)


class VacanciesListAPI(APIView):
    permission_classes = [IsAdminOrReadOnly]
    def get(self, request):
        try:
            vacancies = Vacancies.objects.all()
            serializer = AddVacanciesSerializer(vacancies, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class VacancyApplyAPI(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, id):
        try:
            vacancy = Vacancies.objects.get(id=id)
            if VacanciesMember.objects.filter(vacancy=vacancy, user=request.user).exists():
                return Response({'error': 'Вы уже подали заявку на эту вакансию'}, status=status.HTTP_400_BAD_REQUEST)
            VacanciesMember.objects.create(vacancy=vacancy, user=request.user)
        except Vacancies.DoesNotExist:
            return Response({'error': 'Вакансия не найдена'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class VacancyMembersListAPI(APIView):
    permission_classes = [IsAuthenticated, IsEmployerOrAdminOrSuperAdmin]
    def get(self, request, vacancy_id):
        try:
            vacancy = Vacancies.objects.get(id=vacancy_id)
            user = request.user
            user_profile = getattr(user, 'userprofile', None)
            is_authorized = (
                user.is_superuser or
                (user_profile and user_profile.admin_type in ['super']) or
                (user_profile and user_profile.admin_type == 'employer' and vacancy.company.owner == user)
            )
            if not is_authorized:
                return Response(
                    {"error": "У вас нет прав для просмотра участников"},
                    status=status.HTTP_403_FORBIDDEN
                )

            members = VacanciesMember.objects.filter(vacancy=vacancy).select_related('user')
            members_data = [
                {
                    'id': member.user.id,
                    'username': member.user.username,
                    'email': member.user.email
                }
                for member in members
            ]
            return Response(members_data, status=status.HTTP_200_OK)
        except Vacancies.DoesNotExist:
            return Response(
                {"error": "Вакансия не найдена"},
                status=status.HTTP_404_NOT_FOUND
            )


class VacancyMemberRemoveAPI(APIView):
    permission_classes = [IsAuthenticated, IsEmployerOwner]
    def delete(self, request, id, user_id):
        try:
            vacancy = Vacancies.objects.get(id=id)
            if vacancy.company.owner != request.user:
                return Response({'error': 'У вас нет прав'}, status=status.HTTP_403_FORBIDDEN)
            member = VacanciesMember.objects.get(vacancy=vacancy, user__id=user_id)
            member.delete()
            return Response({'message': 'Пользователь успешно исключён'}, status=status.HTTP_204_NO_CONTENT)
        except Vacancies.DoesNotExist:
            return Response({'error': 'Вакансия не найдена'}, status=status.HTTP_404_NOT_FOUND)
        except VacanciesMember.DoesNotExist:
            return Response({'error': 'Пользователь не является участником этой вакансии'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class MyCompaniesAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            companies = Company.objects.filter(owner=request.user)
            serializer = CompanySerializer(companies, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class MyVacanciesAPI(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            vacancies = Vacancies.objects.filter(company__owner=request.user)
            serializer = AddVacanciesSerializer(vacancies, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

class ApplyVacancyAPI(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, vacancy_id):
        try:
            vacancy = Vacancies.objects.get(id=vacancy_id)
        except Vacancies.DoesNotExist:
            return Response(
                {"error": "Вакансия не найдена"},
                status=status.HTTP_404_NOT_FOUND
            )
        if VacanciesMember.objects.filter(vacancy=vacancy, user=request.user).exists():
            return Response(
                {"error": "Вы уже подали заявку на эту вакансию"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            application = VacanciesMember.objects.create(
                vacancy=vacancy,
                user=request.user
            )
            return Response(
                {"message": "Заявка успешно подана", "application_id": application.id},
                status=status.HTTP_201_CREATED
            )
        except IntegrityError as e:
            return Response(
                {"error": "Не удалось подать заявку: возможно, заявка уже существует"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"Не удалось подать заявку: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


# спасибо за документацию!!!
class VacancyMembersExportAPI(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, vacancy_id):
        try:
            vacancy = Vacancies.objects.get(id=vacancy_id)
        except Vacancies.DoesNotExist:
            return Response(
                {"error": "Вакансия не найдена"},
                status=status.HTTP_404_NOT_FOUND
            )
        user = request.user
        user_profile = getattr(user, 'userprofile', None)
        is_authorized = (
            user.is_superuser or
            (user_profile and user_profile.admin_type in ['super']) or
            (user_profile and user_profile.admin_type == 'employer' and vacancy.company.owner == user)
        )
        if not is_authorized:
            return Response(
                {"error": "У вас нет прав для скачивания списка участников"},
                status=status.HTTP_403_FORBIDDEN
            )

        members = VacanciesMember.objects.filter(vacancy=vacancy).select_related('user', 'user__userprofile')
        if not members.exists():
            return Response(
                {"message": "Нет участников для этой вакансии"},
                status=status.HTTP_200_OK
            )
        wb = Workbook()
        ws = wb.active
        ws.title = f"Участники вакансии {vacancy_id}"

        headers = [
            "ID пользователя",
            "Имя",
            "Фамилия",
            "Отчество",
            "Email",
            "Телефон",
            "Название вакансии",
            "ID вакансии"
        ]
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header
            ws[f"{get_column_letter(col_num)}1"].font = Font(bold=True)

        for row_num, member in enumerate(members, 2):
            user = member.user
            profile = getattr(user, 'userprofile', None)
            ws[f"A{row_num}"] = user.id
            ws[f"B{row_num}"] = profile.name if profile else ''
            ws[f"C{row_num}"] = profile.surname if profile else ''
            ws[f"D{row_num}"] = profile.patronymic if profile else ''
            ws[f"E{row_num}"] = user.email
            ws[f"F{row_num}"] = profile.phone if profile else ''
            ws[f"G{row_num}"] = vacancy.title
            ws[f"H{row_num}"] = vacancy.id

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            content=output.read()
        )
        response['Content-Disposition'] = f'attachment; filename=vacancy_{vacancy_id}_members.xlsx'
        return response